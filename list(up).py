from __future__ import annotations

import io
import math
import re
import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple, Dict, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side

# =========================
# ì•±/í…Œë§ˆ
# =========================
st.set_page_config(page_title="ë¸Œëœë“œë³„ í…œí”Œë¦¿ ë¹Œë”(í´ë¼ìš°ë“œ)", layout="wide")

# =========================
# ìƒìˆ˜/ìŠ¤íƒ€ì¼
# =========================
PRODUCT_NAME_CANDIDATES = ["ë“±ë¡ìƒí’ˆëª…", "ìƒí’ˆëª…", "ë…¸ì¶œìƒí’ˆëª…", "ì¿ íŒ¡ë…¸ì¶œìƒí’ˆëª…"]
OPTION_NAME_COL = "ë“±ë¡ ì˜µì…˜ëª…"

RE_SIZE = re.compile(r"(?<!\d)(\d{3})(?!\d)")
WS = re.compile(r"\s+")
COLOR_LONG2SHORT = {
    "ë¸”ë™": "í‘", "í™”ì´íŠ¸": "ë°±", "ì•„ì´ë³´ë¦¬": "ì•„ì´ë³´ë¦¬",
    "ê·¸ë ˆì´": "íšŒ", "ë² ì´ì§€": "ë² ", "ë„¤ì´ë¹„": "ê³¤",
    "í•‘í¬": "í•‘", "ë ˆë“œ": "ì ", "ë¸Œë¼ìš´": "ë°¤",
    "ì‹¤ë²„": "ì‹¤", "ê³¨ë“œ": "ê¸ˆ",
}
SHORT_COLORS = {"í‘","ë°±","ì•„ì´ë³´ë¦¬","íšŒ","ë² ","ê³¤","í•‘","ì ","ë°¤","ì‹¤","ê¸ˆ"}
COLOR_WORDS = list(COLOR_LONG2SHORT.keys()) + list(SHORT_COLORS)
BRAND_TITLE = {"ì¤€ë””": "ì¤€ë””ìì¸", "ììœ ": "ììœ ë””ìì¸"}

BRIGHT_YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN   = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# =========================
# ê³µí†µ ìœ í‹¸
# =========================
def normalize_text(s: str) -> str:
    if pd.isna(s):
        return ""
    return WS.sub(" ", str(s)).strip()

def find_product_name_col(df: pd.DataFrame) -> Optional[str]:
    for c in PRODUCT_NAME_CANDIDATES:
        if c in df.columns:
            return c
    stripped = {col.replace(" ", ""): col for col in df.columns}
    for c in PRODUCT_NAME_CANDIDATES:
        cc = c.replace(" ", "")
        if cc in stripped:
            return stripped[cc]
    return None

def shorten_color(token: str) -> str:
    if not token:
        return token
    t = str(token).strip()
    return COLOR_LONG2SHORT.get(t, t)

def _today_md() -> str:
    now = dt.datetime.now()
    return f"{now.month}/{now.day}"

def _clean_value(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.lower() in {"none", "nan"}:
        return None
    return v

def _write(ws, row, col, v, as_int=False):
    v = _clean_value(v)
    if v is None:
        ws.cell(row=row, column=col, value=None)
        return
    if as_int:
        try:
            ws.cell(row=row, column=col, value=int(v)); return
        except Exception:
            pass
    ws.cell(row=row, column=col, value=v)

# =========================
# ì˜µì…˜ íŒŒì‹±
# =========================
def _preprocess_option_text(text: str) -> str:
    if not text:
        return ""
    s = normalize_text(text)
    s = re.sub(r'\b(ìƒ‰ìƒ|ì‚¬ì´ì¦ˆ)\b\s*[:ï¼š]?', ' ', s)
    color_alt = "|".join(map(re.escape, COLOR_WORDS))
    s = re.sub(rf'(ì—¬ì„±|ë‚¨ì„±)\s*({color_alt})', r'\1 \2', s)
    s = re.sub(r'(ì—¬ì„±|ë‚¨ì„±)(\d{3})', r'\1 \2', s)
    return WS.sub(" ", s).strip()

def _find_color_near_size(s: str, size_span: Tuple[int,int]) -> Optional[str]:
    color_regex = re.compile("|".join(sorted(map(re.escape, COLOR_WORDS), key=len, reverse=True)))
    candidates = [(m.group(), m.start(), m.end()) for m in color_regex.finditer(s)]
    if not candidates:
        return None
    size_center = (size_span[0] + size_span[1]) / 2
    best = None; best_dist = 10**9
    for word, a, b in candidates:
        center = (a + b) / 2
        dist = abs(center - size_center)
        if dist < best_dist:
            best_dist = dist; best = word
    return best

def parse_color_size(option_name: str) -> Tuple[Optional[str], Optional[str]]:
    if not option_name:
        return None, None
    text = _preprocess_option_text(option_name)
    m = RE_SIZE.search(text)
    if not m:
        return None, None
    size = m.group(1)
    color_word = _find_color_near_size(text, m.span())
    if not color_word:
        toks = text.split()
        for tok in reversed(toks):
            if tok in COLOR_WORDS:
                color_word = tok; break
    if not color_word:
        return None, None
    base_short = shorten_color(color_word)
    is_women = 'ì—¬ì„±' in text; is_men = 'ë‚¨ì„±' in text
    if is_women:   color_label = f"{base_short}ä¸­"
    elif is_men:   color_label = f"{base_short}å¤§"
    else:          color_label = base_short
    return color_label, size

def extract_colors_sizes(option_series: pd.Series) -> Tuple[List[str], List[str]]:
    colors, sizes = [], []; seen_c, seen_s = set(), set()
    for v in option_series.dropna().astype(str):
        c, s = parse_color_size(v)
        if c and c not in seen_c: colors.append(c); seen_c.add(c)
        if s and s not in seen_s: sizes.append(s); seen_s.add(s)
    return colors, sizes

# =========================
# ì„¸ì…˜/ê·¸ë¦¬ë“œ
# =========================
def init_session():
    st.session_state.setdefault("basket", [])
    st.session_state.setdefault("template_bytes", None)
    st.session_state.setdefault("manual_product", "")
    st.session_state.setdefault("manual_colors_text", "")
    st.session_state.setdefault("manual_size_start", 230)
    st.session_state.setdefault("manual_size_end", 280)
    st.session_state.setdefault("manual_size_step", 5)

def add_grid_to_basket(grid_df: pd.DataFrame, brand: str, product: str):
    for color in grid_df.index:
        for size in grid_df.columns:
            val = grid_df.loc[color, size]
            if pd.isna(val): continue
            try:
                q = int(val)
            except:
                continue
            if q <= 0: continue
            st.session_state["basket"].append({
                "brand": brand, "product": product,
                "color": str(color), "size": str(size), "qty": q
            })

def clear_basket():
    st.session_state["basket"] = []

def build_grid(colors: List[str], sizes: List[str], selectable_colors: List[str]) -> pd.DataFrame:
    if not sizes: return pd.DataFrame()
    if not selectable_colors: selectable_colors = []
    df = pd.DataFrame(index=selectable_colors, columns=sizes, data=math.nan)
    df.index.name = "ìƒ‰ìƒ"
    return df

# =========================
# í…œí”Œë¦¿ ì¡°ì‘
# =========================
def _set_header(ws, brand: str):
    header_txt = f"{_today_md()} {BRAND_TITLE.get(brand, brand)}(&Pë²ˆ)"
    try:
        ws.header_footer.center_header = header_txt
    except Exception:
        try:
            ws.oddHeader.center.text = header_txt
            ws.evenHeader.center.text = header_txt
        except Exception:
            pass
    try:
        ws.page_setup.firstPageNumber = 51
        ws.page_setup.use_firstPageNumber = True
    except Exception:
        pass

def _clear_data_area_keep_row1(ws, start_row=2, start_col=2):
    maxr = ws.max_row or 2000
    maxc = ws.max_column or 200
    for r in range(start_row, maxr + 1):
        for c in range(start_col, maxc + 1):
            ws.cell(row=r, column=c).value = None

def export_brand_template_xlsx(brand: str, items: List[dict], template_bytes: bytes) -> bytes:
    """
    - í…œí”Œë¦¿(ì—…ë¡œë“œ/ë ˆí¬ íŒŒì¼)ì˜ 1í–‰/ì„œì‹ ë³´ì¡´
    - B2ë¶€í„° ê¸°ë¡, ì œí’ˆëª… í•˜ì´ë¼ì´íŠ¸(FFFF00)
    - ì œí’ˆë³„ 'ì‹¤ì‚¬ìš© ì‚¬ì´ì¦ˆ'ë§Œ ê°€ë¡œ í—¤ë”(C~) ìƒì„±, Cì—´ ì´í›„ ë„ˆë¹„=4
    """
    try:
        wb = load_workbook(io.BytesIO(template_bytes))
    except Exception:
        wb = Workbook()
    ws = wb.active
    _set_header(ws, brand)
    _clear_data_area_keep_row1(ws, start_row=2, start_col=2)

    by_product: Dict[str, List[dict]] = {}
    for it in items:
        by_product.setdefault(it["product"], []).append(it)

    row = 2
    for product_raw, p_items in by_product.items():
        # ì‹¤ì œ ì‚¬ìš© ì‚¬ì´ì¦ˆ ìˆ˜ì§‘
        sizes_used, seen = [], set()
        for it in p_items:
            try:
                q = int(it["qty"])
            except:
                continue
            if q > 0:
                sz = str(it["size"])
                if sz not in seen:
                    seen.add(sz); sizes_used.append(sz)

        def _size_key(x):
            m = re.search(r"\d+", str(x))
            return (0, int(m.group())) if m else (1, str(x))
        sizes_used.sort(key=_size_key)

        # ìƒ‰ìƒë³„ í•©ê³„
        by_color: Dict[str, Dict[str, int]] = {}
        for it in p_items:
            try:
                q = int(it["qty"])
            except:
                continue
            if q <= 0: continue
            color = str(it["color"]); sz = str(it["size"])
            by_color.setdefault(color, {})
            by_color[color][sz] = by_color[color].get(sz, 0) + q

        # ì œí’ˆëª…
        _write(ws, row, 2, product_raw)
        ws.cell(row=row, column=2).fill = BRIGHT_YELLOW

        # ì‚¬ì´ì¦ˆ í—¤ë”
        start_col = 3
        for i, sz in enumerate(sizes_used):
            col = start_col + i
            _write(ws, row, col, sz)
            ws.column_dimensions[get_column_letter(col)].width = 4

        # ìƒ‰ìƒ í–‰
        for color, size_map in by_color.items():
            row += 1
            _write(ws, row, 2, color)
            for i, sz in enumerate(sizes_used):
                col = start_col + i
                val = size_map.get(sz)
                _write(ws, row, col, val, as_int=True)

        row += 1

    # ë§ˆë¬´ë¦¬: C~ ë§ˆì§€ë§‰ê¹Œì§€ ë„ˆë¹„=4
    maxc = ws.max_column or 3
    for c in range(3, maxc + 1):
        ws.column_dimensions[get_column_letter(c)].width = 4

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# =========================
# ë ˆí¬ íŒŒì¼ ìë™ ì¸ì‹ + ì—…ë¡œë“œ ëŒ€ì²´
# =========================
@st.cache_data(show_spinner=False)
def read_local_excel_sheets(path: Path | None):
    if not path or not path.exists():
        return None
    try:
        xls = pd.ExcelFile(path)
        return {s: pd.read_excel(xls, sheet_name=s) for s in xls.sheet_names}
    except Exception as e:
        st.warning(f"ë ˆí¬ íŒŒì¼ ì½ê¸° ì‹¤íŒ¨({path.name}): {e}")
        return None

def pick_first_present(paths: list[Path]) -> Path | None:
    for p in paths:
        if p.exists():
            return p
    return None

def load_uploaded_df(file) -> Optional[pd.DataFrame]:
    try:
        df = pd.read_excel(file)
        if OPTION_NAME_COL not in df.columns:
            st.error(f"ì—‘ì…€ì— '{OPTION_NAME_COL}' ì»¬ëŸ¼ì´ í•„ìš”í•©ë‹ˆë‹¤.")
            return None
        name_col = find_product_name_col(df)
        if not name_col:
            st.error(f"ì œí’ˆëª… ì»¬ëŸ¼ì´ í•„ìš”í•©ë‹ˆë‹¤. í›„ë³´: {', '.join(PRODUCT_NAME_CANDIDATES)}")
            return None
        df[name_col] = df[name_col].map(normalize_text)
        df[OPTION_NAME_COL] = df[OPTION_NAME_COL].map(normalize_text)
        return df
    except Exception as e:
        st.exception(e); return None

def sheets_from_upload(file):
    if not file:
        return None
    try:
        if file.name.lower().endswith(".csv"):
            return {"Sheet1": pd.read_csv(file)}
        return {s: pd.read_excel(file, sheet_name=s) for s in pd.ExcelFile(file).sheet_names}
    except Exception as e:
        st.error(f"ì—…ë¡œë“œ íŒŒì¼ ì½ê¸° ì‹¤íŒ¨: {e}")
        return None

# =========================
# ê²€ìƒ‰/ê·¸ë¦¬ë“œ ë¹Œë”
# =========================
def search_products(df: pd.DataFrame, keyword: str, name_col: str) -> pd.DataFrame:
    if not keyword:
        return df[[name_col]].drop_duplicates().head(200)
    kw = keyword.strip()
    mask = df[name_col].astype(str).str.contains(re.escape(kw), case=False, na=False)
    return df.loc[mask, [name_col]].drop_duplicates().head(200)

def _sizes_range_step(start: int, end: int, step: int) -> List[str]:
    if step <= 0 or end < start: return []
    return [str(s) for s in range(start, end + 1, step)]

# =========================
# íƒ­ ë³¸ë¬¸
# =========================
def tab_body(brand: str, df: pd.DataFrame):
    name_col = find_product_name_col(df)

    st.subheader(f"{brand} â€” ì œí’ˆëª… ê²€ìƒ‰")
    q = st.text_input("ì œí’ˆëª… ê²€ìƒ‰ (ë¶€ë¶„ ê²€ìƒ‰)", key=f"q_{brand}")
    result = search_products(df, q, name_col)
    st.caption(f"ê²€ìƒ‰ ê²°ê³¼ {len(result)}ê°œ (ìµœëŒ€ 200ê°œ í‘œì‹œ)")

    product = st.selectbox(
        "ì œí’ˆ ì„ íƒ",
        ["(ì„ íƒ)"] + result[name_col].tolist(),
        key=f"product_{brand}"
    )
    manual_toggle = st.checkbox("ì§ì ‘ì…ë ¥ ëª¨ë“œ (íŒŒì‹± ë¬´ì‹œí•˜ê³  ìˆ˜ë™ ì‘ì„±)", key=f"manual_{brand}")

    # ìë™ íŒŒì‹±
    if product != "(ì„ íƒ)" and not manual_toggle:
        sub = df[df[name_col] == product]
        colors, sizes = extract_colors_sizes(sub[OPTION_NAME_COL])

        if not sizes:
            st.warning("ì´ ì œí’ˆì€ 'ë“±ë¡ ì˜µì…˜ëª…'ì—ì„œ ì‚¬ì´ì¦ˆë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. ì•„ë˜ 'ì§ì ‘ì…ë ¥ ëª¨ë“œ'ë¥¼ ì‚¬ìš©í•˜ì„¸ìš”.")
            manual_toggle = True

        if not manual_toggle:
            st.write("---")
            st.subheader("2) í‘œ ì…ë ¥ (ì„¸ë¡œ=ìƒ‰ìƒ / ê°€ë¡œ=ì‚¬ì´ì¦ˆ)")
            color_choices = st.multiselect(
                "ì„¸ë¡œ(í–‰)ì— í‘œì‹œí•  ìƒ‰ìƒ ì„ íƒ",
                options=colors,
                default=colors[: min(5, len(colors))],
                key=f"colors_{brand}"
            )

            grid = build_grid(colors, sizes, color_choices)
            st.caption("ê° ì¹¸ì— ìˆ˜ëŸ‰ì„ ì…ë ¥í•˜ì„¸ìš”. ë¹„ìš°ê±°ë‚˜ 0ì€ ë¯¸ë°˜ì˜ë©ë‹ˆë‹¤.")
            edited = st.data_editor(
                grid, key=f"editor_{brand}",
                use_container_width=True, num_rows="fixed",
            )

            c1, c2 = st.columns(2)
            with c1:
                if st.button("í˜„ì¬ í‘œë¥¼ ë¦¬ìŠ¤íŠ¸ì—…ì— ì¶”ê°€", key=f"add_{brand}"):
                    add_grid_to_basket(edited, brand=brand, product=product)
                    st.success("ë¦¬ìŠ¤íŠ¸ì—…ì— ì¶”ê°€í–ˆìŠµë‹ˆë‹¤.")
            with c2:
                if st.button("ë¦¬ìŠ¤íŠ¸ì—… ë¹„ìš°ê¸°", key=f"clear_{brand}"):
                    clear_basket(); st.info("ë¦¬ìŠ¤íŠ¸ì—…ì„ ë¹„ì› ìŠµë‹ˆë‹¤.")

    # ì§ì ‘ ì…ë ¥
    if manual_toggle:
        st.write("---"); st.subheader("ì§ì ‘ì…ë ¥ ëª¨ë“œ")
        mp = st.text_input("ì œí’ˆëª…(ì§ì ‘ ì…ë ¥)", value=(product if product != "(ì„ íƒ)" else st.session_state.get("manual_product", "")), key=f"mp_{brand}")
        st.session_state["manual_product"] = mp

        st.caption("ìì£¼ ì“°ëŠ” ìƒ‰ìƒ í”„ë¦¬ì…‹")
        preset_colors = ["í‘","ë°±","ê³¤","í•‘","íšŒ","ë² ","ì ","ì•„ì´ë³´ë¦¬"]
        cols = st.columns(len(preset_colors))
        for i, col in enumerate(cols):
            if col.button(preset_colors[i], key=f"preset_{brand}_{i}"):
                mc_text = st.session_state.get("manual_colors_text", "")
                parts = [p.strip() for p in mc_text.split(",") if p.strip()]
                if preset_colors[i] not in parts: parts.append(preset_colors[i])
                st.session_state["manual_colors_text"] = ",".join(parts)

        mc_text = st.text_input(
            "ìƒ‰ìƒ ëª©ë¡(ì‰¼í‘œë¡œ êµ¬ë¶„, ì˜ˆ: í‘, ë°±, ì•„ì´ë³´ë¦¬)",
            value=st.session_state.get("manual_colors_text", ""),
            key=f"mc_{brand}"
        )
        st.session_state["manual_colors_text"] = mc_text
        colors = [c.strip() for c in mc_text.split(",") if c.strip()]

        c1, c2, c3 = st.columns(3)
        with c1:
            start = st.number_input("ì‹œì‘ ì‚¬ì´ì¦ˆ", min_value=200, max_value=400, value=st.session_state.get("manual_size_start", 230), step=5, key=f"start_{brand}")
        with c2:
            end = st.number_input("ë ì‚¬ì´ì¦ˆ", min_value=200, max_value=420, value=st.session_state.get("manual_size_end", 280), step=5, key=f"end_{brand}")
        with c3:
            step = st.number_input("ê°„ê²©", min_value=1, max_value=50, value=st.session_state.get("manual_size_step", 5), step=1, key=f"step_{brand}")

        st.session_state["manual_size_start"] = int(start)
        st.session_state["manual_size_end"] = int(end)
        st.session_state["manual_size_step"] = int(step)

        sizes = _sizes_range_step(int(start), int(end), int(step))

        if not mp:
            st.info("ì œí’ˆëª…ì„ ì…ë ¥í•˜ë©´ í‘œê°€ ìƒì„±ë©ë‹ˆë‹¤."); return
        if not colors:
            st.info("ìƒ‰ìƒì„ í•œ ê°œ ì´ìƒ ì…ë ¥í•˜ì„¸ìš”. (ì‰¼í‘œë¡œ êµ¬ë¶„)"); return
        if not sizes:
            st.warning("ì‚¬ì´ì¦ˆ ë²”ìœ„/ê°„ê²©ì„ ë‹¤ì‹œ í™•ì¸í•˜ì„¸ìš”."); return

        st.subheader(f"2) í‘œ ì…ë ¥ (ì„¸ë¡œ=ìƒ‰ìƒ / ê°€ë¡œ=ì‚¬ì´ì¦ˆ {start}~{end}, {step} ë‹¨ìœ„)")
        grid = build_grid(colors, sizes, colors)
        st.caption("ê° ì¹¸ì— ìˆ˜ëŸ‰ì„ ì…ë ¥í•˜ì„¸ìš”. ë¹„ìš°ê±°ë‚˜ 0ì€ ë¯¸ë°˜ì˜ë©ë‹ˆë‹¤.")
        edited = st.data_editor(
            grid, key=f"editor_manual_{brand}",
            use_container_width=True, num_rows="fixed",
        )

        c1, c2 = st.columns(2)
        with c1:
            if st.button("í˜„ì¬ í‘œë¥¼ ë¦¬ìŠ¤íŠ¸ì—…ì— ì¶”ê°€(ì§ì ‘ì…ë ¥)", key=f"add_manual_{brand}"):
                add_grid_to_basket(edited, brand=brand, product=mp); st.success("ë¦¬ìŠ¤íŠ¸ì—…ì— ì¶”ê°€í–ˆìŠµë‹ˆë‹¤.")
        with c2:
            if st.button("ë¦¬ìŠ¤íŠ¸ì—… ë¹„ìš°ê¸°", key=f"clear_manual_{brand}"):
                clear_basket(); st.info("ë¦¬ìŠ¤íŠ¸ì—…ì„ ë¹„ì› ìŠµë‹ˆë‹¤.")

    # í˜„ì¬ ë¸Œëœë“œë§Œ ë³´ê¸°/ë‹¤ìš´ë¡œë“œ
    basket = [e for e in st.session_state["basket"] if e["brand"] == brand]
    st.write("---"); st.subheader(f"{brand} ëˆ„ì  ë¦¬ìŠ¤íŠ¸ì—… ë¯¸ë¦¬ë³´ê¸°")
    if not basket:
        st.caption("ì•„ì§ ì¶”ê°€ëœ í•­ëª©ì´ ì—†ìŠµë‹ˆë‹¤.")
    else:
        st.dataframe(pd.DataFrame(basket), use_container_width=True, height=280)
        if not st.session_state["template_bytes"]:
            st.error("template.xlsx ë°”ì´íŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤. ì¢Œì¸¡/ë ˆí¬ì—ì„œ í…œí”Œë¦¿ì„ ì œê³µí•´ ì£¼ì„¸ìš”.")
        else:
            xlsx_bytes = export_brand_template_xlsx(brand, basket, st.session_state["template_bytes"])
            fname = f"{brand}_{dt.datetime.now():%Y%m%d}_template.xlsx"
            st.download_button(
                f"{brand} í…œí”Œë¦¿ ë‹¤ìš´ë¡œë“œ (xlsx)",
                data=xlsx_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# =========================
# ë©”ì¸
# =========================
def main():
    init_session()

    st.title("TXT ì—†ì´ ë°”ë¡œ í…œí”Œë¦¿ ë§Œë“¤ê¸° (ë¸Œëœë“œë³„ íŒŒì¼ Â· í´ë¼ìš°ë“œ)")
    st.caption("í…œí”Œë¦¿ ë³´ì¡´ Â· B2ë¶€í„° ê¸°ë¡ Â· ì œí’ˆëª… í•˜ì´ë¼ì´íŠ¸=FFFF00 Â· ì œí’ˆë³„ ê°€ë³€ ì‚¬ì´ì¦ˆ í—¤ë” Â· Cì—´ë¶€í„° ë„ˆë¹„=4 ê³ ì •")

    # --- ë ˆí¬ ê²½ë¡œ ìë™ ì¸ì‹ ---
    base = Path(__file__).parent
    data_dir = base / "data"
    tpl_path   = None
    jundi_path = None
    jayu_path  = None
    if data_dir.exists():
        tpl_path   = pick_first_present([data_dir / "template.xlsx"])
        jundi_path = pick_first_present([data_dir / "jundi.xlsx"])
        jayu_path  = pick_first_present([data_dir / "jayu.xlsx"])

    local_tpl_sheets   = read_local_excel_sheets(tpl_path)    if tpl_path   else None
    local_jundi_sheets = read_local_excel_sheets(jundi_path)  if jundi_path else None
    local_jayu_sheets  = read_local_excel_sheets(jayu_path)   if jayu_path  else None

    # --- ì‚¬ì´ë“œë°” (í˜„í™© + ì—…ë¡œë“œë¡œ ë®ì–´ì“°ê¸°) ---
    with st.sidebar:
        st.header("ğŸ“ ë°ì´í„° ì†ŒìŠ¤")
        def status(label, ok, path: Path | None):
            p = (str(path.relative_to(base)) if path else label)
            return f"âœ… {p}" if ok else f"â›” {p} (ì—†ìŒ)"

        st.caption(status("data/template.xlsx",   bool(local_tpl_sheets),   tpl_path))
        st.caption(status("data/jundi.xlsx",     bool(local_jundi_sheets),  jundi_path))
        st.caption(status("data/jayu.xlsx",      bool(local_jayu_sheets),   jayu_path))

        st.write("---")
        st.write("**ì—…ë¡œë“œë¡œ ëŒ€ì²´/ë³´ì™„**")
        tpl_up   = st.file_uploader("template.xlsx", type=["xlsx"], key="tpl_up")
        jundi_up = st.file_uploader("ì¤€ë”” ìƒí’ˆë¦¬ìŠ¤íŠ¸", type=["xlsx"], key="jundi_up")
        jayu_up  = st.file_uploader("ììœ  ìƒí’ˆë¦¬ìŠ¤íŠ¸", type=["xlsx"], key="jayu_up")

    # ì—…ë¡œë“œ â†’ ì‹œíŠ¸ dict
    tpl_sheets   = sheets_from_upload(tpl_up)   or local_tpl_sheets
    jundi_sheets = sheets_from_upload(jundi_up) or local_jundi_sheets
    jayu_sheets  = sheets_from_upload(jayu_up)  or local_jayu_sheets

    if not tpl_sheets:
        st.warning("template.xlsxê°€ í•„ìš”í•©ë‹ˆë‹¤. ë ˆí¬ data/ì— ì˜¬ë¦¬ê±°ë‚˜ ì—…ë¡œë“œí•˜ì„¸ìš”.")
        st.stop()

    # í…œí”Œë¦¿ ë°”ì´íŠ¸(ë ˆí¬ íŒŒì¼ ìš°ì„ )
    template_bytes = None
    if tpl_path and tpl_path.exists():
        try:
            template_bytes = tpl_path.read_bytes()
        except Exception as e:
            st.warning(f"ë ˆí¬ í…œí”Œë¦¿ ë°”ì´íŠ¸ ì½ê¸° ì‹¤íŒ¨: {e}")
    if not template_bytes:
        # DataFrame dictë¡œ ëŒ€ì²´ í…œí”Œë¦¿ (ì„œì‹ ìµœì†Œ) â€” ê°€ëŠ¥í•˜ë©´ ë ˆí¬ í…œí”Œë¦¿ ì‚¬ìš© ê¶Œì¥
        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "TEMPLATE_PLACEHOLDER"
            bio = io.BytesIO(); wb.save(bio)
            template_bytes = bio.getvalue()
        except Exception:
            st.error("í…œí”Œë¦¿ ë°”ì´íŠ¸ ìƒì„± ì‹¤íŒ¨"); st.stop()
    st.session_state["template_bytes"] = template_bytes

    # ìƒí’ˆë¦¬ìŠ¤íŠ¸ íƒ­ ì¤€ë¹„
    tabs: List[Tuple[str, pd.DataFrame]] = []
    if jundi_sheets:
        # ì²« ì‹œíŠ¸ ì‚¬ìš©
        name = next(iter(jundi_sheets.keys()))
        df = jundi_sheets[name]
        df = load_uploaded_df(io.BytesIO(df.to_excel(index=False)) if isinstance(df, pd.DataFrame) else df) or df
        if isinstance(df, pd.DataFrame) and OPTION_NAME_COL in df.columns:
            tabs.append(("ì¤€ë””", df))
        else:
            # ì—…ë¡œë“œ/ë ˆí¬ê°€ ë°”ë¡œ DataFrameì´ë©´ ê·¸ëŒ€ë¡œ ê²€ì‚¬
            if isinstance(jundi_sheets[name], pd.DataFrame):
                df2 = jundi_sheets[name]
                if OPTION_NAME_COL in df2.columns:
                    tabs.append(("ì¤€ë””", df2))
    if jayu_sheets:
        name = next(iter(jayu_sheets.keys()))
        df = jayu_sheets[name]
        df = load_uploaded_df(io.BytesIO(df.to_excel(index=False)) if isinstance(df, pd.DataFrame) else df) or df
        if isinstance(df, pd.DataFrame) and OPTION_NAME_COL in df.columns:
            tabs.append(("ììœ ", df))
        else:
            if isinstance(jayu_sheets[name], pd.DataFrame):
                df2 = jayu_sheets[name]
                if OPTION_NAME_COL in df2.columns:
                    tabs.append(("ììœ ", df2))

    st.write("---")
    if not tabs:
        st.info("data/jundi.xlsx ë˜ëŠ” data/jayu.xlsx ë¥¼ ì œê³µí•˜ê±°ë‚˜ ì—…ë¡œë“œí•˜ë©´ íƒ­ì´ ë‚˜íƒ€ë‚©ë‹ˆë‹¤.")
        return

    t_objs = st.tabs([t[0] for t in tabs])
    for tab, (brand, df) in zip(t_objs, tabs):
        with tab:
            tab_body(brand, df)

if __name__ == "__main__":
    main()
